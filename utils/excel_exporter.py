"""Export data to Excel format with formatting"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import logging
import os


class ExcelExporter:
    """Export scraped data to Excel with formatting"""
    
    def __init__(self):
        self.logger = logging.getLogger('excel_exporter')
    
    def export_to_excel(self, df, filename, apply_formatting=True):
        """
        Export DataFrame to Excel with optional formatting
        
        Args:
            df: pandas DataFrame
            filename: Output filename
            apply_formatting: Whether to apply Excel formatting
        """
        try:
            self.logger.info(f"Exporting data to {filename}...")
            
            # Ensure output directory exists
            output_dir = os.path.dirname(filename)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            # Export to Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Wheels Data')
                
                if apply_formatting:
                    workbook = writer.book
                    worksheet = writer.sheets['Wheels Data']
                    self._apply_formatting(worksheet, df)
            
            self.logger.info(f"✓ Successfully exported {len(df)} rows to {filename}")
            
            # Print file size
            file_size = os.path.getsize(filename) / (1024 * 1024)  # Convert to MB
            self.logger.info(f"✓ File size: {file_size:.2f} MB")
            
        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {str(e)}")
            raise
    
    def _apply_formatting(self, worksheet, df):
        """
        Apply formatting to Excel worksheet
        
        Args:
            worksheet: openpyxl worksheet
            df: pandas DataFrame
        """
        try:
            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            cell_alignment = Alignment(vertical="top", wrap_text=False)
            border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            # Format header row
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Auto-adjust column widths
            column_widths = {}
            for col_num, column in enumerate(df.columns, 1):
                column_letter = get_column_letter(col_num)
                
                # Calculate max width for this column
                max_length = len(str(column))  # Header length
                
                for idx, value in enumerate(df[column].astype(str), 2):
                    # Limit check to first 100 rows for performance
                    if idx > 100:
                        break
                    max_length = max(max_length, len(value))
                
                # Set column width (with max limit)
                adjusted_width = min(max_length + 2, 50)
                column_widths[column_letter] = adjusted_width
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply borders and alignment to data cells
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border
                    cell.alignment = cell_alignment
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
            # Set specific column widths for known columns
            specific_widths = {
                'url': 40,
                'image_url': 40,
                'date': 18,
                'sku': 15,
                'pn': 15,
                'actual_price': 12,
                'msrp': 12,
                'title': 50,
                'description': 60,
                'year': 8,
                'make': 15,
                'model': 20,
                'trims': 30,
                'engines': 30
            }
            
            for col_num, column in enumerate(df.columns, 1):
                if column in specific_widths:
                    column_letter = get_column_letter(col_num)
                    worksheet.column_dimensions[column_letter].width = specific_widths[column]
            
            self.logger.info("✓ Excel formatting applied")
            
        except Exception as e:
            self.logger.error(f"Error applying formatting: {str(e)}")
    
    def export_summary(self, stats, filename):
        """
        Export summary statistics to a separate sheet or file
        
        Args:
            stats: Dictionary of summary statistics
            filename: Output filename
        """
        try:
            # Create summary DataFrame
            summary_data = []
            for key, value in stats.items():
                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        summary_data.append({
                            'Metric': f"{key} - {sub_key}",
                            'Value': str(sub_value)
                        })
                else:
                    summary_data.append({
                        'Metric': key,
                        'Value': str(value)
                    })
            
            summary_df = pd.DataFrame(summary_data)
            
            # Add summary to existing Excel file or create new one
            if os.path.exists(filename):
                # Add to existing file
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                    summary_df.to_excel(writer, index=False, sheet_name='Summary')
                    self.logger.info("✓ Summary added to existing Excel file")
            else:
                # Create new file with summary
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    summary_df.to_excel(writer, index=False, sheet_name='Summary')
                    self.logger.info("✓ Summary exported to new Excel file")
            
        except Exception as e:
            self.logger.error(f"Error exporting summary: {str(e)}")
    
    def split_by_site(self, df, output_dir):
        """
        Split the main Excel file by website/brand
        
        Args:
            df: pandas DataFrame
            output_dir: Output directory for split files
        """
        try:
            os.makedirs(output_dir, exist_ok=True)
            
            # Extract site name from URL
            df['site'] = df['url'].apply(lambda x: x.split('/')[2] if isinstance(x, str) and len(x.split('/')) > 2 else 'unknown')
            
            # Split by site
            for site, group in df.groupby('site'):
                site_filename = os.path.join(output_dir, f"wheels_{site}.xlsx")
                
                # Remove the temporary 'site' column before export
                group_export = group.drop('site', axis=1)
                
                self.export_to_excel(group_export, site_filename, apply_formatting=True)
                self.logger.info(f"✓ Exported {len(group)} rows for {site}")
            
            # Remove temporary column from original df
            df.drop('site', axis=1, inplace=True)
            
        except Exception as e:
            self.logger.error(f"Error splitting by site: {str(e)}")

